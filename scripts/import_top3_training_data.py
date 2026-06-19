"""从 _top3_analysis xlsx 导入选中的训练数据到对应类别目录。

读取 imagestest/_top3_analysis/battlefield_*/ 下的 xlsx 文件，
将"是否选作训练数据"列标记为1的行对应的裁剪图复制到
training/data/images/{class_id}/ 目录下，按训练集命名规范重命名。

用法:
    python import_top3_training_data.py
"""

import re
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl

PROJECT_ROOT = Path(__file__).parent.parent.parent
TOPDIR = PROJECT_ROOT / "imagestest" / "_top3_analysis"
TRAIN_IMAGES = Path(__file__).parent.parent / "data" / "images"


def extract_class_id(final_str: str) -> int:
    """从'56(骷髅兵)'或'56(骷髅兵) [双]'格式提取class_id。"""
    match = re.match(r"(\d+)", str(final_str).strip())
    if match:
        return int(match.group(1))
    return -1


def get_next_index(class_dir: Path, class_id: int) -> int:
    """获取该类别的下一个可用索引。"""
    if not class_dir.exists():
        class_dir.mkdir(parents=True, exist_ok=True)
        return 0
    existing = list(class_dir.glob(f"{class_id}_*.png"))
    if not existing:
        return 0
    indices = []
    for f in existing:
        try:
            idx = int(f.stem.split("_")[-1])
            indices.append(idx)
        except ValueError:
            continue
    return max(indices) + 1 if indices else 0


def process_xlsx(xlsx_path: Path, crops_dir: Path) -> int:
    """处理单个xlsx文件，返回导入的图片数。"""
    if not xlsx_path.exists():
        print(f"  [跳过] xlsx不存在: {xlsx_path}")
        return 0

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    print(f"  处理: {xlsx_path.name} ({ws.max_row - 1}行数据)")

    imported = 0
    skipped_empty = 0
    skipped_obstacle = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        # col B=图片id(索引1), col I=最终识别结果(索引8), col J=是否选作训练数据(索引9)
        img_id = str(row[1]) if len(row) > 1 and row[1] else None
        final_str = str(row[8]) if len(row) > 8 and row[8] else None
        selected = str(row[9]).strip() if len(row) > 9 and row[9] else ""

        if not img_id or selected != "1":
            continue
        if not final_str:
            continue

        class_id = extract_class_id(final_str)
        if class_id < 0:
            continue

        # 跳过空地(190)和障碍(189) — 这些不是兵种类别
        if class_id == 190:
            skipped_empty += 1
            continue
        if class_id == 189:
            skipped_obstacle += 1
            continue

        # 源裁剪图
        src_crop = crops_dir / f"{int(img_id):04d}.png"
        if not src_crop.exists():
            print(f"    [缺失] {src_crop}")
            continue

        # 目标目录和文件名
        dst_dir = TRAIN_IMAGES / str(class_id)
        dst_dir.mkdir(parents=True, exist_ok=True)
        next_idx = get_next_index(dst_dir, class_id)
        dst_file = dst_dir / f"{class_id}_{next_idx}.png"

        shutil.copy2(src_crop, dst_file)
        imported += 1

    wb.close()
    if skipped_empty or skipped_obstacle:
        print(f"    跳过了 {skipped_empty} 个空地(190) + {skipped_obstacle} 个障碍(189) 标记")
    return imported


def main():
    """主流程: 遍历 _top3_analysis 下的所有 battlefield 目录。"""
    print("=" * 60)
    print("  导入 Top-3 分析训练数据")
    print("=" * 60)

    total = 0
    for bf_dir in sorted(TOPDIR.glob("battlefield_*")):
        if not bf_dir.is_dir():
            continue
        xlsx_file = bf_dir / f"{bf_dir.name}_top3_analysis.xlsx"
        crops_dir = bf_dir / "crops"
        if not xlsx_file.exists():
            continue
        n = process_xlsx(xlsx_file, crops_dir)
        print(f"    导入了 {n} 张图片")
        total += n

    print(f"\n总计导入: {total} 张训练图片")
    print("完成 (未启动训练)")


if __name__ == "__main__":
    main()
