"""生成兵种-图片映射XLSX文件，方便查看每个兵种的图片覆盖情况。

用法:
    cd F:/桌面/test3 && python training/scripts/create_xlsx_mapping.py

输出:
    - training/data/creature_image_mapping.xlsx  兵种-图片映射表
    - gamedata/creature_sparse_vectors.py         稀疏向量工具模块
"""

import json
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

IMAGES_DIR = PROJECT_ROOT / "images" / "creatures"
LABELS_FILE = PROJECT_ROOT / "gamedata" / "creature_labels.json"
OUTPUT_XLSX = PROJECT_ROOT / "training" / "data" / "creature_image_mapping.xlsx"
SPARSE_VECTORS_OUT = PROJECT_ROOT / "gamedata" / "creature_sparse_vectors.py"


def load_labels() -> dict[str, Any]:
    with open(LABELS_FILE, encoding="utf-8") as f:
        return json.load(f)


def _normalize(s: str) -> str:
    return s.lower().replace(" ", "_").replace("-", "_").replace("'", "")


def find_images_for_creature(name_en: str, label_index: int) -> list[str]:
    """查找某个兵种的所有图片文件。

    支持两种命名格式:
    1. 新格式: {label_index}_{index}.png
    2. 旧格式: 包含英文名的文件名
    """
    norm_name = _normalize(name_en)
    images: list[str] = []
    for img in sorted(IMAGES_DIR.glob("*.png")):
        stem = img.stem
        # 新格式匹配: 数字_数字
        import re
        m = re.match(r"^(\d+)_\d+$", stem)
        if m:
            if int(m.group(1)) == label_index:
                images.append(img.name)
            continue
        # 旧格式匹配
        img_norm = _normalize(stem)
        if norm_name in img_norm:
            images.append(img.name)
    return images


def build_mapping() -> list[dict[str, Any]]:
    """构建每个兵种到其图片的映射。"""
    labels_data = load_labels()
    labels = labels_data["labels"]

    rows: list[dict[str, Any]] = []
    for label in labels:
        images = find_images_for_creature(label["name_en"], label["label"])

        creature_prefix_new = f"{label['label']}_0.png"
        has_main = any(img == creature_prefix_new for img in images)

        rows.append({
            "label_index": label["label"],
            "name_en": label["name_en"],
            "name_zh": label.get("name_zh", ""),
            "faction": label.get("faction", ""),
            "level": label.get("level", 0),
            "is_upgraded": "是" if label.get("is_upgraded") else "否",
            "main_image": "有" if has_main else "缺失",
            "extra_images": len(images) - (1 if has_main else 0),
            "total_images": len(images),
            "image_files": ", ".join(images) if images else "无",
        })

    return rows


def create_xlsx(rows: list[dict[str, Any]]) -> None:
    """生成XLSX文件，含嵌入式图片缩略图、样式和筛选。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import io
    except ImportError as e:
        print(f"[WARN] 缺少依赖({e})，生成CSV版本")
        _create_csv_fallback(rows)
        return

    TEMP_DIR = PROJECT_ROOT / "training" / "data" / "_thumbnails"
    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ---- Sheet 1: 主映射表 ----
    ws1 = wb.active
    ws1.title = "兵种图片映射"

    headers = [
        "标签ID", "英文名", "中文名", "阵营", "等级", "已升级",
        "主图状态", "额外图片数", "总图片数", "兵种图片",
    ]
    col_widths = [8, 28, 18, 14, 6, 8, 10, 12, 10, 16]

    # 表头样式
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # 条件填充
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # 预生成缩略图并嵌入
    THUMB_SIZE = (64, 64)
    IMAGES_DIR = PROJECT_ROOT / "images" / "creatures"

    # 数据行
    for row_idx, row_data in enumerate(rows, 2):
        # 文本列
        for col_idx, header in enumerate(headers, 1):
            if header == "兵种图片":
                continue
            value = row_data.get(_header_to_key(header), "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 嵌入图片缩略图 (列10)
        label_idx = row_data["label_index"]
        img_files = find_images_for_creature(row_data["name_en"], label_idx)
        img_column = 10  # "兵种图片" 列索引

        if img_files:
            first_img = img_files[0]
            img_path = IMAGES_DIR / first_img
            if img_path.exists():
                try:
                    pil_img = PILImage.open(img_path).convert("RGBA")
                    pil_img = pil_img.resize(THUMB_SIZE, PILImage.LANCZOS)
                    buf = io.BytesIO()
                    pil_img.save(buf, format="PNG")
                    buf.seek(0)

                    xl_img = XLImage(buf)
                    xl_img.width = 64
                    xl_img.height = 64
                    cell_ref = f"{get_column_letter(img_column)}{row_idx}"
                    ws1.add_image(xl_img, cell_ref)
                except Exception as e:
                    ws1.cell(row=row_idx, column=img_column, value=f"加载失败: {e}")
            else:
                ws1.cell(row=row_idx, column=img_column, value="文件缺失")
        else:
            ws1.cell(row=row_idx, column=img_column, value="无图片")

        # 行高
        ws1.row_dimensions[row_idx].height = 68

        # 着色
        if row_data["total_images"] == 0:
            fill = red_fill
        elif row_data["total_images"] == 1:
            fill = yellow_fill
        else:
            fill = green_fill

        for col_idx in range(1, len(headers) + 1):
            ws1.cell(row=row_idx, column=col_idx).fill = fill

    # 筛选和冻结
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws1.freeze_panes = "A2"

    # ---- Sheet 2: 统计摘要 ----
    ws2 = wb.create_sheet("统计摘要")
    total = len(rows)
    no_img = sum(1 for r in rows if r["total_images"] == 0)
    one_img = sum(1 for r in rows if r["total_images"] == 1)
    multi_img = sum(1 for r in rows if r["total_images"] > 1)

    stats = [
        ("总兵种数", total),
        ("无图片(需下载)", no_img),
        ("仅1张图片(需扩充)", one_img),
        ("多张图片(达标)", multi_img),
        ("图片覆盖率", f"{total - no_img}/{total} ({(total - no_img) / total * 100:.1f}%)"),
    ]

    for i, (label, value) in enumerate(stats, 1):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, size=12)
        ws2.cell(row=i, column=2, value=value).font = Font(size=12)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    # ---- Sheet 3: 按阵营汇总 ----
    ws3 = wb.create_sheet("阵营汇总")
    faction_stats: dict[str, dict[str, int]] = {}
    for r in rows:
        f = r["faction"]
        if f not in faction_stats:
            faction_stats[f] = {"total": 0, "with_img": 0, "missing": 0}
        faction_stats[f]["total"] += 1
        if r["total_images"] > 0:
            faction_stats[f]["with_img"] += 1
        else:
            faction_stats[f]["missing"] += 1

    faction_headers = ["阵营", "总兵种", "有图片", "缺图片", "覆盖率"]
    for col_idx, h in enumerate(faction_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, (faction, stats_data) in enumerate(sorted(faction_stats.items()), 2):
        ws3.cell(row=i, column=1, value=faction).border = thin_border
        ws3.cell(row=i, column=2, value=stats_data["total"]).border = thin_border
        ws3.cell(row=i, column=3, value=stats_data["with_img"]).border = thin_border
        ws3.cell(row=i, column=4, value=stats_data["missing"]).border = thin_border
        coverage = f"{stats_data['with_img'] / stats_data['total'] * 100:.0f}%"
        ws3.cell(row=i, column=5, value=coverage).border = thin_border

    for col_idx in range(1, 6):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 15

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_XLSX))
    print(f"[OK] XLSX已保存: {OUTPUT_XLSX}")
    print(f"  Sheet1: 兵种图片映射 (含嵌入式缩略图, {len(rows)}行)")
    print(f"  Sheet2: 统计摘要")
    print(f"  Sheet3: 阵营汇总")

    # 清理临时缩略图
    import shutil
    if TEMP_DIR.exists():
        shutil.rmtree(TEMP_DIR, ignore_errors=True)


def _header_to_key(header: str) -> str:
    mapping = {
        "标签ID": "label_index", "英文名": "name_en", "中文名": "name_zh",
        "阵营": "faction", "等级": "level", "已升级": "is_upgraded",
        "主图状态": "main_image", "额外图片数": "extra_images",
        "总图片数": "total_images", "兵种图片": "image_thumbnail",
    }
    return mapping.get(header, header)


def _create_csv_fallback(rows: list[dict[str, Any]]) -> None:
    import csv
    csv_path = OUTPUT_XLSX.with_suffix(".csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        if rows:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
    print(f"[OK] CSV已保存: {csv_path}")


def generate_sparse_vectors_module(rows: list[dict[str, Any]]) -> None:
    """生成稀疏向量工具模块。"""
    code = '''"""兵种稀疏向量工具 — 提供171类兵种的one-hot编码。

每个兵种对应一个唯一的类别索引(0-170)，可用于:
- 模型分类输出层 (171类softmax)
- 稀疏向量检索 (RAG知识库)
- 标签转换 (name_en <-> label_index)

用法:
    from gamedata.creature_sparse_vectors import (
        NUM_CLASSES, label_to_vector, vector_to_label, get_label_by_name
    )
    vec = label_to_vector(5)  # -> [0,0,0,0,0,1,0,...,0] (171维)
    name = vector_to_label(vec)  # -> "Billy Goat"
"""

import torch
from typing import Union, Optional
import json
from pathlib import Path

# 自动加载标签文件
_LABELS_PATH = Path(__file__).parent / "creature_labels.json"
with open(_LABELS_PATH, encoding="utf-8") as f:
    _data = json.load(f)

NUM_CLASSES: int = _data["num_classes"]
_LABELS: list[dict] = _data["labels"]

# 建立查找索引
_INDEX_TO_CREATURE: dict[int, dict] = {l["label"]: l for l in _LABELS}
_NAME_TO_INDEX: dict[str, int] = {l["name_en"]: l["label"] for l in _LABELS}
_NAMEZH_TO_INDEX: dict[str, int] = {l["name_zh"]: l["label"] for l in _LABELS}


def label_to_vector(label_index: int) -> list[int]:
    """将标签索引转换为稀疏向量(one-hot编码)。

    Args:
        label_index: 兵种标签索引 (0~170)

    Returns:
        171维one-hot向量

    Raises:
        ValueError: 标签索引超出范围
    """
    if not 0 <= label_index < NUM_CLASSES:
        raise ValueError(f"标签索引 {label_index} 超出范围 [0, {NUM_CLASSES})")
    vec = [0] * NUM_CLASSES
    vec[label_index] = 1
    return vec


def label_to_tensor(label_index: int) -> "torch.Tensor":
    """将标签索引转换为PyTorch张量(one-hot)。

    Args:
        label_index: 兵种标签索引 (0~170)

    Returns:
        shape=(NUM_CLASSES,) 的float32张量
    """
    t = torch.zeros(NUM_CLASSES, dtype=torch.float32)
    t[label_index] = 1.0
    return t


def vector_to_label(vector: Union[list[int], "torch.Tensor"]) -> str:
    """将稀疏向量转换回兵种英文名。

    Args:
        vector: one-hot向量(list或tensor)

    Returns:
        兵种英文名
    """
    if isinstance(vector, torch.Tensor):
        idx = int(vector.argmax().item())
    else:
        idx = vector.index(max(vector))
    creature = _INDEX_TO_CREATURE.get(idx)
    return creature["name_en"] if creature else f"Unknown({idx})"


def get_label_by_name(name: str, lang: str = "en") -> Optional[int]:
    """根据名称查找标签索引。

    Args:
        name: 兵种名称
        lang: 语言, "en"(英文) 或 "zh"(中文)

    Returns:
        标签索引, 未找到返回None
    """
    if lang == "zh":
        return _NAMEZH_TO_INDEX.get(name)
    return _NAME_TO_INDEX.get(name)


def get_creature_info(label_index: int) -> Optional[dict]:
    """根据标签索引获取兵种完整信息。

    Args:
        label_index: 兵种标签索引 (0~170)

    Returns:
        包含name_en, name_zh, faction, level, is_upgraded的字典
    """
    return _INDEX_TO_CREATURE.get(label_index)


def get_labels_by_faction(faction: str) -> list[int]:
    """获取指定阵营的所有标签索引。

    Args:
        faction: 阵营名 (如 "Castle", "Rampart", "Cove" 等)

    Returns:
        标签索引列表
    """
    return [l["label"] for l in _LABELS if l.get("faction") == faction]


def get_all_labels() -> list[dict]:
    """获取所有兵种标签信息。"""
    return _LABELS.copy()


def get_faction_list() -> list[str]:
    """获取所有阵营名列表。"""
    factions = sorted(set(l.get("faction", "") for l in _LABELS))
    return [f for f in factions if f]
'''

    with open(SPARSE_VECTORS_OUT, "w", encoding="utf-8") as f:
        f.write(code)
    print(f"[OK] 稀疏向量模块已生成: {SPARSE_VECTORS_OUT}")


if __name__ == "__main__":
    print("=" * 60)
    print("兵种-图片映射XLSX生成器")
    print("=" * 60)

    rows = build_mapping()

    # 统计
    no_img = [r for r in rows if r["total_images"] == 0]
    one_img = [r for r in rows if r["total_images"] == 1]
    multi_img = [r for r in rows if r["total_images"] > 1]

    print(f"\n总兵种: {len(rows)}")
    print(f"  无图片: {len(no_img)} 个")
    if no_img:
        for r in no_img:
            print(f"    - [{r['label_index']}] {r['name_en']} ({r['name_zh']}) [{r['faction']}]")
    print(f"  仅1张图片: {len(one_img)} 个")
    print(f"  多张图片: {len(multi_img)} 个")

    create_xlsx(rows)
    # 注意: 稀疏向量模块由generate_sparse_vectors_module输出,
    # 但该模块已被手动编辑(含_HAS_TORCH guard等), 请勿自动覆盖。
    # generate_sparse_vectors_module(rows)

    print("\n完成!")
